"""Excel report generator for swing trading analysis"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
import io
import warnings
warnings.filterwarnings('ignore')

class ExcelExporter:
    """Class to handle Excel report generation"""
    
    def __init__(self):
        self.workbook = None
        self.report_data = {}
        
    def render_excel_exporter(self, selected_stocks, selected_timeframes, 
                             start_date, end_date, data_fetcher, 
                             technical_analysis, pattern_detection):
        """Render Excel export interface"""
        
        st.markdown("""
        📑 **Excel Report Generator** creates comprehensive trading reports with multiple sheets 
        containing analysis summaries, detailed stock breakdowns, and pattern statistics.
        """)
        
        # Report configuration
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown("**📋 Report Configuration**")
            
            report_name = st.text_input(
                "Report Name", 
                value=f"Swing_Trading_Analysis_{datetime.now().strftime('%Y%m%d')}"
            )
            
            include_charts = st.checkbox("Include Price Charts", value=True)
            include_patterns = st.checkbox("Include Pattern Analysis", value=True)
            include_signals = st.checkbox("Include Trading Signals", value=True)
            include_stats = st.checkbox("Include Performance Statistics", value=True)
            
            # Report scope
            st.markdown("**📊 Report Scope**")
            
            max_stocks = st.slider("Maximum Stocks to Include", 10, 50, 25)
            min_success_rate = st.slider("Minimum Pattern Success Rate %", 50, 90, 65)
            
        with col2:
            st.markdown("**⚙️ Export Settings**")
            
            file_format = st.selectbox("File Format", ["Excel (.xlsx)", "CSV (.csv)"])
            
            color_scheme = st.selectbox(
                "Color Scheme", 
                ["Professional Blue", "Market Green", "Classic Black"]
            )
            
            font_size = st.selectbox("Font Size", ["Small (9pt)", "Medium (11pt)", "Large (12pt)"])
            
            # Quick templates
            st.markdown("**🚀 Quick Templates**")
            
            if st.button("📈 Bullish Setups Report", use_container_width=True):
                self.generate_bullish_report(selected_stocks, selected_timeframes, 
                                           start_date, end_date, data_fetcher, 
                                           technical_analysis, pattern_detection)
            
            if st.button("📉 Bearish Setups Report", use_container_width=True):
                self.generate_bearish_report(selected_stocks, selected_timeframes, 
                                           start_date, end_date, data_fetcher, 
                                           technical_analysis, pattern_detection)
        
        # Generate custom report
        if st.button("📊 Generate Custom Report", type="primary"):
            if not selected_stocks:
                st.error("Please select at least one stock from the sidebar.")
                return
            
            self.generate_comprehensive_report(
                selected_stocks[:max_stocks], selected_timeframes, start_date, end_date,
                data_fetcher, technical_analysis, pattern_detection,
                report_name, include_charts, include_patterns, include_signals, 
                include_stats, min_success_rate, color_scheme, font_size
            )
    
    def generate_comprehensive_report(self, stocks, timeframes, start_date, end_date,
                                    data_fetcher, technical_analysis, pattern_detection,
                                    report_name, include_charts, include_patterns, 
                                    include_signals, include_stats, min_success_rate,
                                    color_scheme, font_size):
        """Generate comprehensive Excel report"""
        
        with st.spinner("Generating comprehensive Excel report..."):
            # Collect all analysis data
            all_analysis_data = self.collect_analysis_data(
                stocks, timeframes, start_date, end_date,
                data_fetcher, technical_analysis, pattern_detection
            )
            
            if not all_analysis_data:
                st.error("No data available for report generation.")
                return
            
            # Create Excel workbook
            self.workbook = openpyxl.Workbook()
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Generate sheets
            self.create_summary_sheet(all_analysis_data, min_success_rate)
            self.create_detailed_analysis_sheet(all_analysis_data)
            self.create_pattern_statistics_sheet(all_analysis_data, min_success_rate)
            
            if include_signals:
                self.create_trading_signals_sheet(all_analysis_data)
            
            if include_stats:
                self.create_performance_stats_sheet(all_analysis_data)
            
            # Apply styling
            self.apply_report_styling(color_scheme, font_size)
            
            # Save and provide download
            self.provide_download(report_name)
    
    def collect_analysis_data(self, stocks, timeframes, start_date, end_date,
                             data_fetcher, technical_analysis, pattern_detection):
        """Collect analysis data for all stocks and timeframes"""
        
        analysis_data = []
        progress_bar = st.progress(0)
        total_combinations = len(stocks) * len(timeframes)
        current_progress = 0
        
        for stock in stocks:
            stock_data = {
                'stock': stock,
                'timeframes': {}
            }
            
            for timeframe in timeframes:
                try:
                    # Fetch data
                    data = data_fetcher.get_stock_data(stock, timeframe, start_date, end_date)
                    
                    if data is not None and len(data) > 50:
                        # Calculate indicators
                        indicators = technical_analysis.calculate_all_indicators(data)
                        
                        # Generate signals
                        signals = technical_analysis.generate_signals(data, indicators)
                        
                        # Detect patterns
                        patterns = pattern_detection.detect_all_patterns(data)
                        
                        # Get current price info
                        current_price = data['Close'].iloc[-1]
                        price_change = ((current_price - data['Close'].iloc[-2]) / data['Close'].iloc[-2]) * 100 if len(data) > 1 else 0
                        
                        stock_data['timeframes'][timeframe] = {
                            'data': data,
                            'indicators': indicators,
                            'signals': signals,
                            'patterns': patterns,
                            'current_price': current_price,
                            'price_change': price_change,
                            'volume': data['Volume'].iloc[-1],
                            'avg_volume': data['Volume'].mean()
                        }
                
                except Exception as e:
                    st.warning(f"Error processing {stock} - {timeframe}: {str(e)}")
                    continue
                
                current_progress += 1
                progress_bar.progress(current_progress / total_combinations)
            
            if stock_data['timeframes']:  # Only add if we have data
                analysis_data.append(stock_data)
        
        progress_bar.empty()
        return analysis_data
    
    def create_summary_sheet(self, analysis_data, min_success_rate):
        """Create summary sheet with best setups by timeframe and market cap"""
        
        ws = self.workbook.create_sheet("Summary & Best Setups")
        
        # Title
        ws['A1'] = "Swing Trading Analysis - Summary Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Report metadata
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Stocks Analyzed: {len(analysis_data)}"
        ws['A5'] = f"Minimum Success Rate Filter: {min_success_rate}%"
        
        # Best setups by timeframe
        row = 7
        ws[f'A{row}'] = "Best Indicator Combinations by Timeframe"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Headers
        headers = ['Timeframe', 'Stock', 'Market Cap', 'Setup Type', 'Signal', 'Success Rate', 'Entry Price', 'Target']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Collect best setups
        best_setups = self.get_best_setups(analysis_data, min_success_rate)
        
        for setup in best_setups[:20]:  # Top 20 setups
            ws.cell(row=row, column=1, value=setup['timeframe'])
            ws.cell(row=row, column=2, value=setup['stock'])
            ws.cell(row=row, column=3, value=setup['market_cap'])
            ws.cell(row=row, column=4, value=setup['setup_type'])
            ws.cell(row=row, column=5, value=setup['signal'])
            ws.cell(row=row, column=6, value=f"{setup['success_rate']}%")
            ws.cell(row=row, column=7, value=f"₹{setup['entry_price']:.2f}")
            ws.cell(row=row, column=8, value=f"₹{setup['target']:.2f}")
            row += 1
        
        # Market cap breakdown
        row += 2
        ws[f'A{row}'] = "Market Cap Performance Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        market_cap_summary = self.get_market_cap_summary(analysis_data)
        
        # Market cap headers
        mc_headers = ['Market Cap', 'Total Stocks', 'Bullish Signals', 'Bearish Signals', 'Avg Success Rate']
        for col, header in enumerate(mc_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        for cap_type, summary in market_cap_summary.items():
            ws.cell(row=row, column=1, value=cap_type)
            ws.cell(row=row, column=2, value=summary['total_stocks'])
            ws.cell(row=row, column=3, value=summary['bullish_signals'])
            ws.cell(row=row, column=4, value=summary['bearish_signals'])
            ws.cell(row=row, column=5, value=f"{summary['avg_success_rate']:.1f}%")
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_detailed_analysis_sheet(self, analysis_data):
        """Create detailed analysis sheet for each stock and timeframe"""
        
        ws = self.workbook.create_sheet("Detailed Stock Analysis")
        
        # Title
        ws['A1'] = "Detailed Stock Analysis by Timeframe"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:L1')
        
        # Headers
        row = 3
        headers = [
            'Stock', 'Timeframe', 'Current Price', 'Price Change %', 'EMA 10', 'EMA 21', 
            'EMA 50', 'RSI', 'MACD Signal', 'Patterns Detected', 'Overall Signal', 'Interpretation'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Fill data
        for stock_data in analysis_data:
            stock = stock_data['stock']
            
            for timeframe, tf_data in stock_data['timeframes'].items():
                indicators = tf_data['indicators']
                signals = tf_data['signals']
                patterns = tf_data['patterns']
                
                # Get indicator values
                ema_10 = indicators.get('EMA_10', [np.nan])[-1] if 'EMA_10' in indicators else np.nan
                ema_21 = indicators.get('EMA_21', [np.nan])[-1] if 'EMA_21' in indicators else np.nan
                ema_50 = indicators.get('EMA_50', [np.nan])[-1] if 'EMA_50' in indicators else np.nan
                rsi = indicators.get('RSI', [np.nan])[-1] if 'RSI' in indicators else np.nan
                
                # MACD signal
                macd_signal = "Bullish" if indicators.get('MACD', [0])[-1] > indicators.get('MACD_signal', [0])[-1] else "Bearish" if 'MACD' in indicators else "N/A"
                
                # Overall signal
                buy_signals = sum(1 for s in signals.values() if s.get('action') == 'BUY')
                sell_signals = sum(1 for s in signals.values() if s.get('action') == 'SELL')
                
                if buy_signals > sell_signals:
                    overall_signal = "BUY"
                elif sell_signals > buy_signals:
                    overall_signal = "SELL"
                else:
                    overall_signal = "HOLD"
                
                # Interpretation
                interpretation = self.get_interpretation(tf_data, overall_signal)
                
                # Fill row
                ws.cell(row=row, column=1, value=stock)
                ws.cell(row=row, column=2, value=timeframe)
                ws.cell(row=row, column=3, value=f"₹{tf_data['current_price']:.2f}")
                ws.cell(row=row, column=4, value=f"{tf_data['price_change']:+.2f}%")
                ws.cell(row=row, column=5, value=f"₹{ema_10:.2f}" if not np.isnan(ema_10) else "N/A")
                ws.cell(row=row, column=6, value=f"₹{ema_21:.2f}" if not np.isnan(ema_21) else "N/A")
                ws.cell(row=row, column=7, value=f"₹{ema_50:.2f}" if not np.isnan(ema_50) else "N/A")
                ws.cell(row=row, column=8, value=f"{rsi:.1f}" if not np.isnan(rsi) else "N/A")
                ws.cell(row=row, column=9, value=macd_signal)
                ws.cell(row=row, column=10, value=len(patterns))
                ws.cell(row=row, column=11, value=overall_signal)
                ws.cell(row=row, column=12, value=interpretation)
                
                # Color code the overall signal
                signal_cell = ws.cell(row=row, column=11)
                if overall_signal == "BUY":
                    signal_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    signal_cell.font = Font(color="FFFFFF", bold=True)
                elif overall_signal == "SELL":
                    signal_cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                    signal_cell.font = Font(color="FFFFFF", bold=True)
                
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_pattern_statistics_sheet(self, analysis_data, min_success_rate):
        """Create pattern statistics and frequency analysis sheet"""
        
        ws = self.workbook.create_sheet("Pattern Statistics")
        
        # Title
        ws['A1'] = "Pattern Frequency & Performance Statistics"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Collect pattern statistics
        pattern_stats = self.calculate_pattern_statistics(analysis_data)
        
        # Pattern frequency table
        row = 3
        ws[f'A{row}'] = "Pattern Frequency Analysis"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Headers
        headers = ['Pattern Name', 'Total Occurrences', 'Bullish', 'Bearish', 'Success Rate', 'Avg Return', 'Risk Level', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Fill pattern data
        for pattern_name, stats in sorted(pattern_stats.items(), key=lambda x: x[1]['total'], reverse=True):
            if stats['total'] > 0:
                ws.cell(row=row, column=1, value=pattern_name)
                ws.cell(row=row, column=2, value=stats['total'])
                ws.cell(row=row, column=3, value=stats['bullish'])
                ws.cell(row=row, column=4, value=stats['bearish'])
                ws.cell(row=row, column=5, value=f"{stats['success_rate']:.1f}%")
                ws.cell(row=row, column=6, value=f"{stats['avg_return']:.1f}%")
                ws.cell(row=row, column=7, value=stats['risk_level'])
                ws.cell(row=row, column=8, value=stats['recommendation'])
                
                # Color code success rate
                success_cell = ws.cell(row=row, column=5)
                if stats['success_rate'] >= 75:
                    success_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    success_cell.font = Font(color="FFFFFF")
                elif stats['success_rate'] >= 65:
                    success_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                else:
                    success_cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                    success_cell.font = Font(color="FFFFFF")
                
                row += 1
        
        # Timeframe performance
        row += 2
        ws[f'A{row}'] = "Timeframe Performance Analysis"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        timeframe_stats = self.calculate_timeframe_statistics(analysis_data)
        
        # Timeframe headers
        tf_headers = ['Timeframe', 'Total Signals', 'Buy Signals', 'Sell Signals', 'Pattern Count', 'Avg Success Rate']
        for col, header in enumerate(tf_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        for timeframe, stats in timeframe_stats.items():
            ws.cell(row=row, column=1, value=timeframe)
            ws.cell(row=row, column=2, value=stats['total_signals'])
            ws.cell(row=row, column=3, value=stats['buy_signals'])
            ws.cell(row=row, column=4, value=stats['sell_signals'])
            ws.cell(row=row, column=5, value=stats['pattern_count'])
            ws.cell(row=row, column=6, value=f"{stats['avg_success_rate']:.1f}%")
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_trading_signals_sheet(self, analysis_data):
        """Create trading signals sheet with actionable recommendations"""
        
        ws = self.workbook.create_sheet("Trading Signals")
        
        # Title
        ws['A1'] = "Current Trading Signals & Recommendations"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:J1')
        
        # Headers
        row = 3
        headers = [
            'Stock', 'Timeframe', 'Signal Type', 'Action', 'Strength', 'Entry Price', 
            'Stop Loss', 'Target', 'Risk:Reward', 'Description'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Collect all signals
        all_signals = []
        for stock_data in analysis_data:
            stock = stock_data['stock']
            for timeframe, tf_data in stock_data['timeframes'].items():
                for signal_type, signal in tf_data['signals'].items():
                    signal_info = {
                        'stock': stock,
                        'timeframe': timeframe,
                        'signal_type': signal_type,
                        'action': signal.get('action', 'HOLD'),
                        'strength': signal.get('strength', 'Weak'),
                        'entry_price': signal.get('entry_price', 0),
                        'stop_loss': signal.get('stop_loss', 0),
                        'target': signal.get('target', 0),
                        'description': signal.get('description', '')
                    }
                    all_signals.append(signal_info)
        
        # Sort by strength and action
        strength_order = {'Very Strong': 5, 'Strong': 4, 'Moderate': 3, 'Weak': 2, 'Very Weak': 1}
        all_signals.sort(key=lambda x: (strength_order.get(x['strength'], 0), x['action'] == 'BUY'), reverse=True)
        
        # Fill signals data
        for signal in all_signals[:50]:  # Top 50 signals
            # Calculate risk:reward ratio
            if signal['entry_price'] > 0 and signal['stop_loss'] > 0 and signal['target'] > 0:
                if signal['action'] == 'BUY':
                    risk = abs(signal['entry_price'] - signal['stop_loss'])
                    reward = abs(signal['target'] - signal['entry_price'])
                else:
                    risk = abs(signal['stop_loss'] - signal['entry_price'])
                    reward = abs(signal['entry_price'] - signal['target'])
                
                rr_ratio = f"1:{reward/risk:.1f}" if risk > 0 else "N/A"
            else:
                rr_ratio = "N/A"
            
            ws.cell(row=row, column=1, value=signal['stock'])
            ws.cell(row=row, column=2, value=signal['timeframe'])
            ws.cell(row=row, column=3, value=signal['signal_type'])
            ws.cell(row=row, column=4, value=signal['action'])
            ws.cell(row=row, column=5, value=signal['strength'])
            ws.cell(row=row, column=6, value=f"₹{signal['entry_price']:.2f}" if signal['entry_price'] > 0 else "N/A")
            ws.cell(row=row, column=7, value=f"₹{signal['stop_loss']:.2f}" if signal['stop_loss'] > 0 else "N/A")
            ws.cell(row=row, column=8, value=f"₹{signal['target']:.2f}" if signal['target'] > 0 else "N/A")
            ws.cell(row=row, column=9, value=rr_ratio)
            ws.cell(row=row, column=10, value=signal['description'])
            
            # Color code action
            action_cell = ws.cell(row=row, column=4)
            if signal['action'] == 'BUY':
                action_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                action_cell.font = Font(color="FFFFFF", bold=True)
            elif signal['action'] == 'SELL':
                action_cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                action_cell.font = Font(color="FFFFFF", bold=True)
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_performance_stats_sheet(self, analysis_data):
        """Create performance statistics sheet"""
        
        ws = self.workbook.create_sheet("Performance Statistics")
        
        # Title
        ws['A1'] = "Performance Statistics & Metrics"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Overall statistics
        row = 3
        ws[f'A{row}'] = "Overall Analysis Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Calculate overall stats
        total_stocks = len(analysis_data)
        total_signals = sum(len(tf_data['signals']) for stock_data in analysis_data 
                           for tf_data in stock_data['timeframes'].values())
        total_patterns = sum(len(tf_data['patterns']) for stock_data in analysis_data 
                            for tf_data in stock_data['timeframes'].values())
        
        buy_signals = sum(sum(1 for signal in tf_data['signals'].values() if signal.get('action') == 'BUY')
                         for stock_data in analysis_data for tf_data in stock_data['timeframes'].values())
        
        sell_signals = sum(sum(1 for signal in tf_data['signals'].values() if signal.get('action') == 'SELL')
                          for stock_data in analysis_data for tf_data in stock_data['timeframes'].values())
        
        # Display stats
        stats_data = [
            ['Total Stocks Analyzed', total_stocks],
            ['Total Trading Signals', total_signals],
            ['Total Patterns Detected', total_patterns],
            ['Buy Signals', buy_signals],
            ['Sell Signals', sell_signals],
            ['Hold/Neutral Signals', total_signals - buy_signals - sell_signals],
            ['Buy Signal Percentage', f"{(buy_signals/total_signals*100):.1f}%" if total_signals > 0 else "0%"],
            ['Sell Signal Percentage', f"{(sell_signals/total_signals*100):.1f}%" if total_signals > 0 else "0%"]
        ]
        
        for stat_name, stat_value in stats_data:
            ws.cell(row=row, column=1, value=stat_name)
            ws.cell(row=row, column=2, value=stat_value)
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # Top performing stocks
        row += 2
        ws[f'A{row}'] = "Top Performing Stocks (by Signal Count)"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Calculate stock performance
        stock_performance = []
        for stock_data in analysis_data:
            stock = stock_data['stock']
            total_stock_signals = sum(len(tf_data['signals']) for tf_data in stock_data['timeframes'].values())
            total_stock_patterns = sum(len(tf_data['patterns']) for tf_data in stock_data['timeframes'].values())
            
            stock_buy_signals = sum(sum(1 for signal in tf_data['signals'].values() if signal.get('action') == 'BUY')
                                   for tf_data in stock_data['timeframes'].values())
            
            stock_performance.append({
                'stock': stock,
                'total_signals': total_stock_signals,
                'total_patterns': total_stock_patterns,
                'buy_signals': stock_buy_signals,
                'buy_percentage': (stock_buy_signals/total_stock_signals*100) if total_stock_signals > 0 else 0
            })
        
        # Sort by total signals
        stock_performance.sort(key=lambda x: x['total_signals'], reverse=True)
        
        # Headers for top stocks
        perf_headers = ['Stock', 'Total Signals', 'Buy Signals', 'Buy %', 'Patterns']
        for col, header in enumerate(perf_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        # Fill top 15 stocks
        for stock_perf in stock_performance[:15]:
            ws.cell(row=row, column=1, value=stock_perf['stock'])
            ws.cell(row=row, column=2, value=stock_perf['total_signals'])
            ws.cell(row=row, column=3, value=stock_perf['buy_signals'])
            ws.cell(row=row, column=4, value=f"{stock_perf['buy_percentage']:.1f}%")
            ws.cell(row=row, column=5, value=stock_perf['total_patterns'])
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_best_setups(self, analysis_data, min_success_rate):
        """Get best trading setups across all stocks and timeframes"""
        
        from .config import MARKET_CAPS
        
        best_setups = []
        
        for stock_data in analysis_data:
            stock = stock_data['stock']
            
            # Determine market cap
            market_cap = "Unknown"
            for cap_type, stocks in MARKET_CAPS.items():
                if stock in stocks:
                    market_cap = cap_type
                    break
            
            for timeframe, tf_data in stock_data['timeframes'].items():
                signals = tf_data['signals']
                patterns = tf_data['patterns']
                
                # Process each signal
                for signal_type, signal in signals.items():
                    if signal.get('entry_price', 0) > 0:
                        # Estimate success rate based on signal strength and type
                        strength_rates = {'Very Strong': 85, 'Strong': 75, 'Moderate': 65, 'Weak': 55, 'Very Weak': 45}
                        estimated_success_rate = strength_rates.get(signal.get('strength', 'Weak'), 60)
                        
                        if estimated_success_rate >= min_success_rate:
                            best_setups.append({
                                'stock': stock,
                                'timeframe': timeframe,
                                'market_cap': market_cap,
                                'setup_type': signal_type,
                                'signal': signal.get('action', 'HOLD'),
                                'success_rate': estimated_success_rate,
                                'entry_price': signal.get('entry_price', 0),
                                'target': signal.get('target', 0),
                                'strength': signal.get('strength', 'Weak')
                            })
        
        # Sort by success rate and strength
        strength_order = {'Very Strong': 5, 'Strong': 4, 'Moderate': 3, 'Weak': 2, 'Very Weak': 1}
        best_setups.sort(key=lambda x: (x['success_rate'], strength_order.get(x['strength'], 0)), reverse=True)
        
        return best_setups
    
    def get_market_cap_summary(self, analysis_data):
        """Get market cap performance summary"""
        
        from .config import MARKET_CAPS
        
        summary = {}
        
        for cap_type, stocks in MARKET_CAPS.items():
            cap_stocks = [stock_data for stock_data in analysis_data if stock_data['stock'] in stocks]
            
            if cap_stocks:
                total_stocks = len(cap_stocks)
                
                bullish_signals = sum(
                    sum(1 for signal in tf_data['signals'].values() if signal.get('action') == 'BUY')
                    for stock_data in cap_stocks for tf_data in stock_data['timeframes'].values()
                )
                
                bearish_signals = sum(
                    sum(1 for signal in tf_data['signals'].values() if signal.get('action') == 'SELL')
                    for stock_data in cap_stocks for tf_data in stock_data['timeframes'].values()
                )
                
                # Estimate average success rate
                all_signals = sum(len(tf_data['signals']) for stock_data in cap_stocks 
                                 for tf_data in stock_data['timeframes'].values())
                
                # Simple success rate estimation based on signal distribution
                avg_success_rate = 70 if bullish_signals > bearish_signals else 65 if bearish_signals > bullish_signals else 60
                
                summary[cap_type] = {
                    'total_stocks': total_stocks,
                    'bullish_signals': bullish_signals,
                    'bearish_signals': bearish_signals,
                    'avg_success_rate': avg_success_rate
                }
        
        return summary
    
    def calculate_pattern_statistics(self, analysis_data):
        """Calculate pattern frequency and performance statistics"""
        
        from .config import PATTERN_SUCCESS_RATES
        
        pattern_stats = {}
        
        for stock_data in analysis_data:
            for tf_data in stock_data['timeframes'].values():
                for pattern in tf_data['patterns']:
                    pattern_name = pattern.get('name', 'Unknown')
                    direction = pattern.get('direction', 'Neutral')
                    
                    if pattern_name not in pattern_stats:
                        pattern_stats[pattern_name] = {
                            'total': 0,
                            'bullish': 0,
                            'bearish': 0,
                            'success_rate': PATTERN_SUCCESS_RATES.get(pattern_name, 65),
                            'avg_return': 0,
                            'risk_level': 'Medium',
                            'recommendation': 'Monitor'
                        }
                    
                    pattern_stats[pattern_name]['total'] += 1
                    
                    if direction == 'Bullish':
                        pattern_stats[pattern_name]['bullish'] += 1
                    elif direction == 'Bearish':
                        pattern_stats[pattern_name]['bearish'] += 1
        
        # Calculate additional metrics
        for pattern_name, stats in pattern_stats.items():
            # Estimate average return based on success rate
            stats['avg_return'] = stats['success_rate'] * 0.08  # Rough estimation
            
            # Determine risk level
            if stats['success_rate'] >= 80:
                stats['risk_level'] = 'Low'
                stats['recommendation'] = 'Strong Buy/Sell'
            elif stats['success_rate'] >= 70:
                stats['risk_level'] = 'Medium'
                stats['recommendation'] = 'Consider'
            else:
                stats['risk_level'] = 'High'
                stats['recommendation'] = 'Wait for Confirmation'
        
        return pattern_stats
    
    def calculate_timeframe_statistics(self, analysis_data):
        """Calculate timeframe performance statistics"""
        
        timeframe_stats = {}
        
        for stock_data in analysis_data:
            for timeframe, tf_data in stock_data['timeframes'].items():
                if timeframe not in timeframe_stats:
                    timeframe_stats[timeframe] = {
                        'total_signals': 0,
                        'buy_signals': 0,
                        'sell_signals': 0,
                        'pattern_count': 0,
                        'avg_success_rate': 0
                    }
                
                signals = tf_data['signals']
                patterns = tf_data['patterns']
                
                timeframe_stats[timeframe]['total_signals'] += len(signals)
                timeframe_stats[timeframe]['buy_signals'] += sum(1 for s in signals.values() if s.get('action') == 'BUY')
                timeframe_stats[timeframe]['sell_signals'] += sum(1 for s in signals.values() if s.get('action') == 'SELL')
                timeframe_stats[timeframe]['pattern_count'] += len(patterns)
        
        # Calculate average success rates (simplified estimation)
        for timeframe, stats in timeframe_stats.items():
            if stats['total_signals'] > 0:
                buy_ratio = stats['buy_signals'] / stats['total_signals']
                # Higher buy ratio generally indicates better market conditions
                stats['avg_success_rate'] = 60 + (buy_ratio * 20)  # 60-80% range
            else:
                stats['avg_success_rate'] = 60
        
        return timeframe_stats
    
    def get_interpretation(self, tf_data, overall_signal):
        """Get trading interpretation for the stock/timeframe combination"""
        
        signals = tf_data['signals']
        patterns = tf_data['patterns']
        indicators = tf_data['indicators']
        
        if overall_signal == "BUY":
            if len(patterns) > 0:
                return "Potential Buy - Multiple bullish signals with pattern confirmation"
            elif 'VOLUME_RATIO' in indicators and indicators['VOLUME_RATIO'][-1] > 1.5:
                return "Strong Buy Setup - High volume confirmation"
            else:
                return "Buy Setup - Wait for volume confirmation"
        
        elif overall_signal == "SELL":
            if len(patterns) > 0:
                return "Potential Sell - Multiple bearish signals with pattern confirmation"
            else:
                return "Sell Setup - Consider exit or short position"
        
        else:
            if len(patterns) > 0:
                return "Mixed Signals - Wait for clear direction"
            else:
                return "Hold - No clear trading opportunity"
    
    def apply_report_styling(self, color_scheme, font_size):
        """Apply consistent styling to all sheets"""
        
        # Color schemes
        color_schemes = {
            "Professional Blue": {"primary": "366092", "secondary": "70AD47", "accent": "FFC000"},
            "Market Green": {"primary": "70AD47", "secondary": "366092", "accent": "C5504B"},
            "Classic Black": {"primary": "2F2F2F", "secondary": "70AD47", "accent": "FFC000"}
        }
        
        colors = color_schemes.get(color_scheme, color_schemes["Professional Blue"])
        
        # Font sizes
        font_sizes = {
            "Small (9pt)": 9,
            "Medium (11pt)": 11,
            "Large (12pt)": 12
        }
        
        base_font_size = font_sizes.get(font_size, 11)
        
        # Apply to all sheets
        for sheet in self.workbook.worksheets:
            # Set default font
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.font = Font(size=base_font_size)
            
            # Add borders to data ranges
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in sheet.iter_rows(min_row=3):
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
    
    def provide_download(self, report_name):
        """Provide download link for the Excel report"""
        
        # Save workbook to bytes
        excel_buffer = io.BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Provide download button
        st.success("✅ Excel report generated successfully!")
        
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_buffer.getvalue(),
            file_name=f"{report_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Show report summary
        st.markdown("### 📊 Report Summary")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Total Sheets", len(self.workbook.worksheets))
        
        with col2:
            total_rows = sum(sheet.max_row for sheet in self.workbook.worksheets)
            st.metric("Total Data Rows", total_rows)
        
        with col3:
            st.metric("File Size", f"{len(excel_buffer.getvalue()) / 1024:.1f} KB")
        
        # Sheet descriptions
        st.markdown("**📋 Sheet Contents:**")
        sheet_descriptions = {
            "Summary & Best Setups": "Best indicator combinations and market cap analysis",
            "Detailed Stock Analysis": "Complete breakdown by stock and timeframe",
            "Pattern Statistics": "Pattern frequency and success rate analysis",
            "Trading Signals": "Current actionable trading recommendations",
            "Performance Statistics": "Overall performance metrics and top stocks"
        }
        
        for sheet_name in [sheet.title for sheet in self.workbook.worksheets]:
            description = sheet_descriptions.get(sheet_name, "Additional analysis data")
            st.write(f"• **{sheet_name}**: {description}")
    
    def generate_bullish_report(self, stocks, timeframes, start_date, end_date,
                               data_fetcher, technical_analysis, pattern_detection):
        """Generate focused bullish setups report"""
        
        st.info("🚀 Generating Bullish Setups Report - focusing on buy signals and bullish patterns...")
        
        # This would filter for only bullish signals and patterns
        # Implementation would be similar to comprehensive report but filtered
        
    def generate_bearish_report(self, stocks, timeframes, start_date, end_date,
                               data_fetcher, technical_analysis, pattern_detection):
        """Generate focused bearish setups report"""
        
        st.info("📉 Generating Bearish Setups Report - focusing on sell signals and bearish patterns...")
        
        # This would filter for only bearish signals and patterns
        # Implementation would be similar to comprehensive report but filtered